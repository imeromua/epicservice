import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def format_stock_report(filepath: str):
    """
    Applies custom formatting to the stock report Excel file.
    - Freezes the first row
    - Auto-fits column widths
    - Colors the header row
    - Applies dashed borders to all cells
    - Bold/Italic fonts for specific columns
    - Applies conditional formatting to 'Місяців без руху' and 'Сума залишку (грн)'
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Dashed border style
        dashed_border = Border(
            left=Side(style='dashed'),
            right=Side(style='dashed'),
            top=Side(style='dashed'),
            bottom=Side(style='dashed')
        )

        # Header color (vibrant indigo/blue)
        header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = dashed_border

        # Find column indices
        col_indices = {}
        for col_idx, cell in enumerate(ws[1], 1):
            val = str(cell.value).strip() if cell.value else ""
            if val in ["Артикул", "Залишок (кількість)", "Місяців без руху", "Сума залишку (грн)"]:
                col_indices[val] = col_idx

        article_col_idx = col_indices.get("Артикул")
        qty_col_idx = col_indices.get("Залишок (кількість)")
        months_col_idx = col_indices.get("Місяців без руху")
        sum_col_idx = col_indices.get("Сума залишку (грн)")

        # Conditional formatting colors and fonts
        yellow_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid") # 6-7 months
        orange_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid") # 7-8 months
        red_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")    # 8+ months / sum > 3000
        
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        bold_italic_font = Font(bold=True, italic=True)
        white_bold_italic_font = Font(color="FFFFFF", bold=True, italic=True)

        # Apply formatting to all rows
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
                
                # Apply column-specific fonts and fills
                if col == article_col_idx:
                    cell.font = bold_font
                elif col == qty_col_idx:
                    cell.font = bold_font
                elif col == sum_col_idx:
                    cell.font = bold_italic_font
                    val = cell.value
                    if isinstance(val, (int, float)) and val > 3000:
                        cell.fill = red_fill
                        cell.font = white_bold_italic_font
                elif col == months_col_idx:
                    cell.font = bold_font
                    val = cell.value
                    if isinstance(val, (int, float)):
                        if 6 <= val < 7:
                            cell.fill = yellow_fill
                        elif 7 <= val < 8:
                            cell.fill = orange_fill
                        elif val >= 8:
                            cell.fill = red_fill
                            cell.font = white_font

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        longest_line = max(len(line) for line in lines)
                        if longest_line > max_length:
                            max_length = longest_line
                except:
                    pass
            
            # Add some padding and set a sensible maximum width
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(filepath)
        wb.close()
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Failed to format Excel file: {e}", exc_info=True)
