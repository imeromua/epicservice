from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
import uvicorn
import sys
import os
import traceback
from aiogram import Bot
from aiogram.types import FSInputFile
import openpyxl
import zipfile
from io import BytesIO
from datetime import datetime, timedelta

# Додаємо шлях до кореневої папки проекту
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from database.engine import async_session
from database.orm import (
    orm_find_products, 
    orm_get_temp_list, 
    orm_add_item_to_temp_list,
    orm_update_temp_list_item_quantity,
    orm_delete_temp_list_item,
    orm_clear_temp_list,
    orm_get_temp_list_department
)
from utils.list_processor import process_and_save_list
from utils.archive_manager import get_user_archives as get_archives_for_user, ACTIVE_DIR, parse_filename
from sqlalchemy.exc import SQLAlchemyError
from config import BOT_TOKEN

app = FastAPI()

# Статичні файли і шаблони
app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))

# Bot instance
bot = Bot(token=BOT_TOKEN)


class SearchRequest(BaseModel):
    query: str
    user_id: int


class AddToListRequest(BaseModel):
    user_id: int
    product_id: int
    quantity: int


class UpdateQuantityRequest(BaseModel):
    user_id: int
    product_id: int
    quantity: int


class DeleteItemRequest(BaseModel):
    user_id: int
    product_id: int


@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    """Повертає 204 No Content щоб не була 404 в логах"""
    return Response(status_code=204)


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Головна сторінка Mini App"""
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/health")
async def health_check():
    """Перевірка стану API"""
    return {"status": "ok"}


@app.get("/api/list/department/{user_id}")
async def get_user_list_department(user_id: int):
    """Отримати відділ поточного списку користувача."""
    try:
        department = await orm_get_temp_list_department(user_id)
        return JSONResponse(content={"department": department}, status_code=200)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.post("/api/search")
async def search_products(req: SearchRequest):
    """
    Пошук товарів за артикулом або назвою.
    Повертає список товарів з детальною інформацією.
    """
    try:
        print(f"🔍 Search request: query='{req.query}', user_id={req.user_id}")
        
        print(f"📞 Calling orm_find_products...")
        products = await orm_find_products(req.query)
        print(f"✅ orm_find_products returned {len(products) if products else 0} products")
        
        if not products:
            print(f"⚠️ No products found")
            return JSONResponse(content={"products": [], "message": "Нічого не знайдено"}, status_code=200)
        
        # Отримуємо temp_list користувача для підрахунку резерву
        async with async_session() as session:
            temp_list = await orm_get_temp_list(req.user_id, session=session)
            user_reserved = {item.product_id: item.quantity for item in temp_list} if temp_list else {}
        
        # Отримуємо відділ поточного списку
        current_department = await orm_get_temp_list_department(req.user_id)
        
        # Формуємо відповідь з детальною інформацією
        result = []
        for product in products:
            try:
                total_quantity = float(product.кількість)
            except (ValueError, TypeError):
                total_quantity = 0.0
            
            available = total_quantity - product.відкладено
            
            user_reserved_qty = user_reserved.get(product.id, 0)
            user_reserved_sum = user_reserved_qty * float(product.ціна)
            
            # Перевіряємо чи товар з іншого відділу
            is_different_department = False
            if current_department is not None and product.відділ != current_department:
                is_different_department = True
            
            result.append({
                "id": product.id,
                "article": product.артикул,
                "name": product.назва,
                "price": float(product.ціна),
                "available": available,
                "department": product.відділ,
                "group": product.група,
                "months_without_movement": product.місяці_без_руху or 0,
                "balance_sum": float(product.сума_залишку or 0.0),
                "reserved": product.відкладено,
                "user_reserved": user_reserved_qty,
                "user_reserved_sum": user_reserved_sum,
                "is_different_department": is_different_department,
                "current_list_department": current_department
            })
        
        print(f"✅ Returning {len(result)} products (current_department={current_department})")
        return JSONResponse(content={"products": result}, status_code=200)
        
    except SQLAlchemyError as e:
        print(f"❌ SQLAlchemy ERROR: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка бази даних", "details": str(e)}, status_code=500)
    except Exception as e:
        print(f"❌ ERROR: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Неочікувана помилка", "details": str(e)}, status_code=500)


@app.get("/api/list/{user_id}")
async def get_user_list(user_id: int):
    """Отримати поточний список товарів користувача."""
    try:
        temp_list = await orm_get_temp_list(user_id)
        if not temp_list:
            return JSONResponse(content={"items": [], "total": 0}, status_code=200)
        items = []
        total_sum = 0.0
        for item in temp_list:
            item_total = float(item.product.ціна) * item.quantity
            total_sum += item_total
            items.append({
                "product_id": item.product.id,
                "article": item.product.артикул,
                "name": item.product.назва,
                "quantity": item.quantity,
                "price": float(item.product.ціна),
                "total": item_total
            })
        return JSONResponse(content={"items": items, "total": total_sum, "count": len(items)}, status_code=200)
    except Exception as e:
        return JSONResponse(content={"error": "Помилка отримання списку", "details": str(e)}, status_code=500)


@app.get("/api/archives/{user_id}")
async def get_user_archives(user_id: int):
    """Отримати список архівних файлів користувача."""
    try:
        print(f"📁 Archives request for user_id={user_id}")
        archives = get_archives_for_user(user_id)
        if not archives:
            return JSONResponse(content={"archives": []}, status_code=200)
        result = []
        for filename, timestamp in archives:
            is_surplus = filename.startswith("лишки_")
            result.append({
                "filename": filename,
                "date": timestamp.strftime("%d.%m.%Y %H:%M"),
                "timestamp": timestamp.isoformat(),
                "is_surplus": is_surplus,
                "type": "Лишки" if is_surplus else "Основний список"
            })
        print(f"✅ Returning {len(result)} archives")
        return JSONResponse(content={"archives": result}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in get_user_archives: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка отримання архівів", "details": str(e)}, status_code=500)


@app.get("/api/statistics/{user_id}")
async def get_user_statistics(user_id: int):
    """Отримати статистику користувача: кількість списків, загальна сума, популярні відділи."""
    try:
        archives = get_archives_for_user(user_id)
        
        if not archives:
            return JSONResponse(content={
                "total_lists": 0,
                "total_amount": 0.0,
                "total_items": 0,
                "popular_department": None,
                "this_month_lists": 0,
                "this_month_amount": 0.0
            }, status_code=200)
        
        total_lists = len(archives)
        total_amount = 0.0
        total_items = 0
        departments = {}
        
        # Дата місяць тому
        month_ago = datetime.now() - timedelta(days=30)
        this_month_lists = 0
        this_month_amount = 0.0
        
        for filename, timestamp in archives:
            file_path = os.path.join(ACTIVE_DIR, filename)
            if not os.path.exists(file_path):
                continue
            
            try:
                # Парсимо Excel для підрахунку
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                
                file_amount = 0.0
                file_items = 0
                
                # Рахуємо товари та суму (пропускаємо заголовок)
                # Новий формат: Артикул, Кількість, Ціна, Сума
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    # Пропускаємо рядки підсумків
                    if str(row[0]).strip() in ["", "К-ть артикулів:", "Зібрано на суму:"]:
                        continue
                    
                    file_items += 1
                    
                    # Сума у 4-й колонці (індекс 3)
                    try:
                        if len(row) > 3 and row[3]:
                            # row[3] може бути числом або рядком "123.45 грн"
                            sum_value = row[3]
                            if isinstance(sum_value, str):
                                sum_value = sum_value.replace(' грн', '').replace(',', '.').strip()
                            file_amount += float(sum_value)
                        else:
                            # Fallback: кількість * ціна (колонки 1 і 2)
                            qty = float(row[1]) if len(row) > 1 and row[1] else 0
                            price = float(row[2]) if len(row) > 2 and row[2] else 0
                            file_amount += qty * price
                    except (ValueError, TypeError, IndexError):
                        pass
                
                wb.close()
                
                total_amount += file_amount
                total_items += file_items
                
                # Статистика за місяць
                if timestamp >= month_ago:
                    this_month_lists += 1
                    this_month_amount += file_amount
                
                # Відділ з імені файлу
                parsed = parse_filename(filename)
                if parsed and "department" in parsed:
                    dept = parsed["department"]
                    departments[dept] = departments.get(dept, 0) + 1
                    
            except Exception as e:
                print(f"⚠️ Error parsing {filename}: {e}")
                continue
        
        # Найпопулярніший відділ
        popular_department = max(departments, key=departments.get) if departments else None
        
        print(f"📊 Stats for user {user_id}: {total_lists} lists, {total_amount:.2f} грн, dept: {popular_department}")
        
        return JSONResponse(content={
            "total_lists": total_lists,
            "total_amount": round(total_amount, 2),
            "total_items": total_items,
            "popular_department": popular_department,
            "this_month_lists": this_month_lists,
            "this_month_amount": round(this_month_amount, 2)
        }, status_code=200)
        
    except Exception as e:
        print(f"❌ ERROR in get_user_statistics: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка отримання статистики"}, status_code=500)


@app.get("/api/archives/download-all/{user_id}")
async def download_all_archives(user_id: int):
    """Завантажити всі архіви користувача як ZIP."""
    try:
        archives = get_archives_for_user(user_id)
        
        if not archives:
            raise HTTPException(status_code=404, detail="No archives found")
        
        # Створюємо ZIP в пам'яті
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, timestamp in archives:
                file_path = os.path.join(ACTIVE_DIR, filename)
                if os.path.exists(file_path):
                    # Додаємо файл до ZIP
                    zip_file.write(file_path, filename)
        
        zip_buffer.seek(0)
        
        # Ім'я ZIP файлу
        zip_filename = f"epicservice_archives_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        
        print(f"📦 Created ZIP with {len(archives)} files for user {user_id}")
        
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={zip_filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in download_all_archives: {type(e).__name__}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Download error")


@app.get("/api/archive/stats/{filename}")
async def get_archive_stats(filename: str, user_id: int):
    """
    Отримати статистику з Excel файлу архіву.
    Парсить файл і повертає: кількість товарів, відділ (з імені файлу), автор (user_id).
    """
    try:
        # Перевірка безпеки
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        # Перевіряємо що файл належить користувачу
        parsed = parse_filename(filename)
        if not parsed or parsed["user_id"] != user_id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        file_path = os.path.join(ACTIVE_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Парсимо Excel
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        items_count = 0
        
        # Пропускаємо заголовок (перший рядок) і рахуємо товари
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:  # Пропускаємо порожні рядки
                continue
            # Пропускаємо рядки підсумків
            if str(row[0]).strip() in ["", "К-ть артикулів:", "Зібрано на суму:"]:
                continue
            items_count += 1
        
        wb.close()
        
        # Отримуємо відділ з імені файлу
        department = parsed.get("department", "Невідомо")
        
        print(f"📊 Stats for {filename}: {items_count} items, department={department}, author={user_id}")
        
        return JSONResponse(content={
            "success": True,
            "items_count": items_count,
            "department": str(department),
            "author_id": user_id
        }, status_code=200)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in get_archive_stats: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={
            "success": False,
            "error": "Помилка читання статистики"
        }, status_code=500)


@app.get("/api/archive/download/{filename}")
async def download_archive(filename: str):
    """Завантажити архівний файл."""
    try:
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        file_path = os.path.join(ACTIVE_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        print(f"📥 Download request: {filename}")
        return FileResponse(path=file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in download_archive: {type(e).__name__}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Download error")


@app.delete("/api/archive/delete/{filename}")
async def delete_archive(filename: str, user_id: int):
    """
    Видалити архівний файл.
    Перевіряє що файл належить користувачу перед видаленням.
    """
    try:
        # Перевірка безпеки: заборонити шляхи з '..' та '/'
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        # Перевіряємо що файл належить користувачу
        parsed = parse_filename(filename)
        if not parsed or parsed["user_id"] != user_id:
            print(f"⚠️ User {user_id} tried to delete file not owned by them: {filename}")
            raise HTTPException(status_code=403, detail="Access denied")
        
        file_path = os.path.join(ACTIVE_DIR, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Видаляємо файл
        os.remove(file_path)
        print(f"🗑️ Deleted archive: {filename} by user {user_id}")
        
        return JSONResponse(content={
            "success": True,
            "message": "Файл видалено"
        }, status_code=200)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in delete_archive: {type(e).__name__}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Delete error")


@app.post("/api/add")
async def add_to_list(req: AddToListRequest):
    """Додати товар до списку."""
    try:
        print(f"➕ Add to list: user_id={req.user_id}, product_id={req.product_id}, quantity={req.quantity}")
        await orm_add_item_to_temp_list(user_id=req.user_id, product_id=req.product_id, quantity=req.quantity)
        print(f"✅ Successfully added to temp list")
        return JSONResponse(content={"success": True, "message": f"Додано {req.quantity} шт."}, status_code=200)
    except ValueError as e:
        # Помилка валідації відділу
        print(f"⚠️ Validation error: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=400)
    except Exception as e:
        print(f"❌ ERROR in add_to_list: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка додавання", "details": str(e)}, status_code=500)


@app.post("/api/update")
async def update_item_quantity(req: UpdateQuantityRequest):
    """Оновити кількість товару."""
    try:
        if req.quantity < 1:
            return JSONResponse(content={"success": False, "message": "Кількість має бути більше 0"}, status_code=400)
        await orm_update_temp_list_item_quantity(user_id=req.user_id, product_id=req.product_id, new_quantity=req.quantity)
        return JSONResponse(content={"success": True, "message": f"Кількість оновлено: {req.quantity} шт."}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in update_item_quantity: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка оновлення", "details": str(e)}, status_code=500)


@app.post("/api/delete")
async def delete_item(req: DeleteItemRequest):
    """Видалити товар зі списку."""
    try:
        await orm_delete_temp_list_item(user_id=req.user_id, product_id=req.product_id)
        return JSONResponse(content={"success": True, "message": "Товар видалено"}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in delete_item: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка видалення", "details": str(e)}, status_code=500)


@app.post("/api/clear/{user_id}")
async def clear_list(user_id: int):
    """Очистити список."""
    try:
        await orm_clear_temp_list(user_id)
        return JSONResponse(content={"success": True, "message": "Список очищено"}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in clear_list: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка очищення", "details": str(e)}, status_code=500)


@app.post("/api/save/{user_id}")
async def save_list_to_excel(user_id: int):
    """
    Зберегти список в Excel.
    WebApp: НЕ відправляє в Telegram, тільки зберігає в archives/active/.
    Файл доступний через вкладку "Архів".
    """
    try:
        print(f"💾 Save list request for user_id={user_id} (webapp - archive only)")
        async with async_session() as session:
            async with session.begin():
                main_list_path, surplus_list_path = await process_and_save_list(session, user_id)
        if not main_list_path and not surplus_list_path:
            print(f"⚠️ List is empty for user {user_id}")
            return JSONResponse(content={"success": False, "message": "Список порожній"}, status_code=400)
        print(f"✅ Files saved: main={main_list_path}, surplus={surplus_list_path}")
        return JSONResponse(content={
            "success": True,
            "message": "✅ Список збережено!",
            "cleared": True,
            "has_main": bool(main_list_path),
            "has_surplus": bool(surplus_list_path)
        }, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in save_list_to_excel: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка збереження списку", "details": str(e)}, status_code=500)


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)
