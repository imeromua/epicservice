# webapp/routers/client.py
"""
Клієнтський роутер для обслуговування користувачів Telegram Mini App.
Містить ендпоїнти для пошуку товарів, управління списками та архівами.

Security hardening status (second pass — TMA initData auth applied):
- SECURED (TMA X-Telegram-Init-Data header required, user identity server-derived):
    POST /search                        — user identity from TMA, not client body
    POST /add, POST /update, POST /delete — user identity from TMA, not client body
    GET /list/{user_id}                 — path user_id ignored; TMA identity used
    GET /list/department/{user_id}      — path user_id ignored; TMA identity used
    POST /save/{user_id}               — path user_id ignored; TMA identity used
    POST /clear/{user_id}              — path user_id ignored; TMA identity used
    GET /archives/{user_id}            — path user_id ignored; TMA identity used
    GET /archives/download-all/{user_id} — path user_id ignored; TMA identity used
    GET /statistics/{user_id}          — path user_id ignored; TMA identity used
    GET /archive/stats/{filename}      — query user_id ignored; TMA identity used
    GET /user/role                     — TMA identity used
    POST /products/filter              — body user_id ignored; TMA identity used
- SECURED (JWT Bearer OR TMA initData, ownership check):
    GET /archive/download/{filename}
    DELETE /archive/delete/{filename}
- JWT-only (standalone / mobile app):
    GET /products/search, GET /products/{article}
    GET /saved-lists, POST /saved-lists, GET /saved-lists/{id}/download
"""

import os
import traceback
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import openpyxl
from aiogram import Bot
from fastapi import APIRouter, Depends, Header, HTTPException
from fastapi.responses import FileResponse, JSONResponse, Response
from pydantic import BaseModel
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import select, func, Float, cast

from config import BOT_TOKEN
from database.engine import async_session
from database.models import Product, SavedList, SavedListItem
from database.orm import (
    orm_add_item_to_temp_list,
    orm_clear_temp_list,
    orm_delete_temp_list_item,
    orm_find_products,
    orm_get_temp_list,
    orm_get_temp_list_department,
    orm_get_user_by_id,
    orm_update_temp_list_item_quantity,
)
from utils.archive_manager import ACTIVE_DIR, get_user_archives as fetch_user_archives, parse_filename
from utils.list_processor import process_and_save_list
from webapp.deps import get_current_user_id, get_current_user_id_any_auth, get_tma_user_id
from webapp.utils.file_safety import is_path_within

router = APIRouter()
bot = Bot(token=BOT_TOKEN)

# Resolved ACTIVE_DIR path for archive path-traversal guards.
_ACTIVE_DIR_PATH = Path(ACTIVE_DIR).resolve()


# === Pydantic Models ===

class SearchRequest(BaseModel):
    query: str
    user_id: int
    offset: int = 0
    limit: int = 500


class AddToListRequest(BaseModel):
    user_id: int
    product_id: int
    quantity: int


class UpdateQuantityRequest(BaseModel):
    user_id: int
    product_id: int
    quantity: int


class DeleteItemRequest(BaseModel):
    user_id: int
    product_id: int


class FilterProductsRequest(BaseModel):
    user_id: int
    departments: List[str] = []  # ["10", "20", "310"]
    sort_by: str = "balance_sum"  # balance_sum, months_without_movement, quantity, article
    offset: int = 0
    limit: int = 500


# === Ендпоїнти ===

@router.get("/user/role")
async def get_user_role(tma_user_id: int = Depends(get_tma_user_id)):
    """Повертає роль користувача з бази даних (TMA auth)."""
    user = await orm_get_user_by_id(tma_user_id)
    if not user:
        return JSONResponse(content={"role": "user"})
    return JSONResponse(content={"role": user.role or "user"})


@router.post("/search")
async def search_products(req: SearchRequest, tma_user_id: int = Depends(get_tma_user_id)):
    """
    Пошук товарів за артикулом або назвою з підтримкою пагінації.
    Повертає список товарів з детальною інформацією + має has_more для infinite scroll.
    User identity is derived from validated TMA initData; req.user_id is ignored.
    """
    try:
        print(f"🔍 Search request: query='{req.query}', tma_user_id={tma_user_id}, offset={req.offset}, limit={req.limit}")
        
        all_products = await orm_find_products(req.query)
        print(f"✅ orm_find_products returned {len(all_products) if all_products else 0} total products")
        
        if not all_products:
            print(f"⚠️ No products found")
            return JSONResponse(content={"products": [], "has_more": False, "total": 0}, status_code=200)
        
        # Пагінація
        total_count = len(all_products)
        products = all_products[req.offset:req.offset + req.limit]
        has_more = (req.offset + req.limit) < total_count
        
        # Отримуємо temp_list користувача для підрахунку резерву
        async with async_session() as session:
            temp_list = await orm_get_temp_list(tma_user_id, session=session)
            user_reserved = {item.product_id: item.quantity for item in temp_list} if temp_list else {}
        
        # Отримуємо відділ поточного списку
        current_department = await orm_get_temp_list_department(tma_user_id)
        
        # Формуємо відповідь з детальною інформацією
        result = []
        for product in products:
            try:
                total_quantity = float(product.кількість)
            except (ValueError, TypeError):
                total_quantity = 0.0
            
            # Отримуємо резерв користувача
            user_reserved_qty = user_reserved.get(product.id, 0)

            # Доступна кількість = загальна - загальний резерв - резерв користувача
            available = total_quantity - product.відкладено - user_reserved_qty

            user_reserved_sum = user_reserved_qty * float(product.ціна)
            
            # Перевіряємо чи товар з іншого відділу
            is_different_department = False
            if current_department is not None and product.відділ != current_department:
                is_different_department = True
            
            result.append({
                "id": product.id,
                "article": product.артикул,
                "name": product.назва,
                "price": float(product.ціна),
                "available": available,
                "department": product.відділ,
                "group": product.група,
                "months_without_movement": product.місяці_без_руху or 0,
                "balance_sum": float(product.сума_залишку or 0.0),
                "reserved": product.відкладено,
                "user_reserved": user_reserved_qty,
                "user_reserved_sum": user_reserved_sum,
                "is_different_department": is_different_department,
                "current_list_department": current_department
            })
        
        print(f"✅ Returning {len(result)} products (offset={req.offset}, has_more={has_more}, total={total_count})")
        return JSONResponse(content={
            "products": result, 
            "has_more": has_more,
            "total": total_count,
            "offset": req.offset,
            "limit": req.limit
        }, status_code=200)
        
    except SQLAlchemyError as e:
        print(f"❌ SQLAlchemy ERROR: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка бази даних", "details": str(e)}, status_code=500)
    except Exception as e:
        print(f"❌ ERROR: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Неочікувана помилка", "details": str(e)}, status_code=500)


@router.post("/products/filter")
async def filter_products(req: FilterProductsRequest, tma_user_id: int = Depends(get_tma_user_id)):
    """
    Фільтрація товарів за відділами з сортуванням та пагінацією.
    Повертає список товарів + статистику по фільтру.
    User identity is derived from validated TMA initData; req.user_id is ignored.
    """
    try:
        print(f"🎛️ Filter request: tma_user_id={tma_user_id}, departments={req.departments}, sort_by={req.sort_by}, offset={req.offset}, limit={req.limit}")
        
        async with async_session() as session:
            # Базовий запит - рахуємо тільки товари, де є ДОСТУПНИЙ залишок (кількість - відкладено > 0)
            query = select(Product).where(
                Product.активний == True,
                (cast(Product.кількість, Float) - func.coalesce(cast(Product.відкладено, Float), 0.0)) > 0
            )
            
            # Фільтр по відділах (якщо вказано) - конвертуємо рядки в числа
            if req.departments:
                dept_integers = [int(d) for d in req.departments]
                query = query.where(Product.відділ.in_(dept_integers))
            
            # Підрахунок загальної кількості (для статистики)
            count_query = select(func.count()).select_from(query.subquery())
            total_count_result = await session.execute(count_query)
            total_count = total_count_result.scalar()
            
            # Сортування
            if req.sort_by == "balance_sum":
                query = query.order_by(Product.сума_залишку.desc())
            elif req.sort_by == "months_without_movement":
                query = query.order_by(Product.місяці_без_руху.desc())
            elif req.sort_by == "quantity":
                query = query.order_by(Product.кількість.desc())
            elif req.sort_by == "article":
                query = query.order_by(Product.артикул.asc())
            else:
                query = query.order_by(Product.сума_залишку.desc())
            
            # Пагінація
            query = query.offset(req.offset).limit(req.limit)
            
            # Виконуємо запит
            result = await session.execute(query)
            products = result.scalars().all()
            
            # Отримуємо temp_list користувача для резерву (TMA identity)
            temp_list = await orm_get_temp_list(tma_user_id, session=session)
            user_reserved = {item.product_id: item.quantity for item in temp_list} if temp_list else {}
            
            # Отримуємо відділ поточного списку
            current_department = await orm_get_temp_list_department(tma_user_id)
            
            # Статистика по фільтру
            stats_query = select(
                func.count(Product.id).label('total_articles'),
                func.sum(Product.сума_залишку).label('total_sum'),
                func.sum(cast(Product.кількість, Float)).label('total_quantity')
            ).where(
                Product.активний == True,
                (cast(Product.кількість, Float) - func.coalesce(cast(Product.відкладено, Float), 0.0)) > 0
            )
            
            # Фільтр по відділах для статистики
            if req.departments:
                stats_query = stats_query.where(Product.відділ.in_(dept_integers))
            
            stats_result = await session.execute(stats_query)
            stats = stats_result.first()
            
            # Формуємо відповідь
            result_products = []
            for product in products:
                try:
                    total_quantity = float(product.кількість)
                except (ValueError, TypeError):
                    total_quantity = 0.0
                
                # Отримуємо резерв користувача
                user_reserved_qty = user_reserved.get(product.id, 0)

                # Доступна кількість = загальна - загальний резерв - резерв користувача
                available = total_quantity - product.відкладено - user_reserved_qty

                user_reserved_sum = user_reserved_qty * float(product.ціна)
                
                # Перевіряємо чи товар з іншого відділу
                is_different_department = False
                if current_department is not None and product.відділ != current_department:
                    is_different_department = True
                
                result_products.append({
                    "id": product.id,
                    "article": product.артикул,
                    "name": product.назва,
                    "price": float(product.ціна),
                    "available": available,
                    "department": product.відділ,
                    "group": product.група,
                    "months_without_movement": product.місяці_без_руху or 0,
                    "balance_sum": float(product.сума_залишку or 0.0),
                    "reserved": product.відкладено,
                    "user_reserved": user_reserved_qty,
                    "user_reserved_sum": user_reserved_sum,
                    "is_different_department": is_different_department,
                    "current_list_department": current_department
                })
            
            # Розраховуємо has_more
            has_more = (req.offset + len(result_products)) < total_count
            
            print(f"✅ Filter returned {len(result_products)} products (total={total_count}, has_more={has_more})")
            
            return JSONResponse(content={
                "products": result_products,
                "has_more": has_more,  # ❗️ Додано на верхній рівень
                "total": total_count,  # ❗️ Додано на верхній рівень
                "offset": req.offset,
                "limit": req.limit,
                "statistics": {
                    "total_articles": stats.total_articles or 0,
                    "total_sum": float(stats.total_sum or 0.0),
                    "total_quantity": float(stats.total_quantity or 0.0),
                    "current_count": len(result_products)
                }
            }, status_code=200)
            
    except SQLAlchemyError as e:
        print(f"❌ SQLAlchemy ERROR: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка бази даних", "details": str(e)}, status_code=500)
    except Exception as e:
        print(f"❌ ERROR in filter_products: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Неочікувана помилка", "details": str(e)}, status_code=500)


@router.get("/products/departments")
async def get_departments():
    """
    Отримати список всіх доступних відділів з кількістю товарів.
    """
    try:
        async with async_session() as session:
            # Виключаємо відділ 0 та рахуємо тільки товари, де доступна кількість > 0
            query = select(
                Product.відділ,
                func.count(Product.id).label('count')
            ).where(
                Product.активний == True,
                (cast(Product.кількість, Float) - func.coalesce(cast(Product.відкладено, Float), 0.0)) > 0,
                Product.відділ != 0,  # виключаємо відділ 0
                Product.відділ.isnot(None)  # виключаємо NULL
            ).group_by(Product.відділ).order_by(Product.відділ)
            
            result = await session.execute(query)
            departments = result.all()
            
            dept_list = [
                {"department": dept.відділ, "count": dept.count}
                for dept in departments
            ]
            
            print(f"📊 Returning {len(dept_list)} departments (filtered out dept 0 and fully reserved items)")
            return JSONResponse(content={"departments": dept_list}, status_code=200)
            
    except Exception as e:
        print(f"❌ ERROR in get_departments: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка отримання відділів"}, status_code=500)


@router.get("/list/{user_id}")
async def get_user_list(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """Отримати поточний список товарів користувача.
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        temp_list = await orm_get_temp_list(tma_user_id)
        if not temp_list:
            return JSONResponse(content={"items": [], "total": 0}, status_code=200)
        items = []
        total_sum = 0.0
        for item in temp_list:
            item_total = float(item.product.ціна) * item.quantity
            total_sum += item_total
            items.append({
                "product_id": item.product.id,
                "article": item.product.артикул,
                "name": item.product.назва,
                "quantity": item.quantity,
                "price": float(item.product.ціна),
                "total": item_total
            })
        return JSONResponse(content={"items": items, "total": total_sum, "count": len(items)}, status_code=200)
    except Exception as e:
        return JSONResponse(content={"error": "Помилка отримання списку", "details": str(e)}, status_code=500)


@router.get("/list/department/{user_id}")
async def get_user_list_department(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """Отримати відділ поточного списку користувача.
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        department = await orm_get_temp_list_department(tma_user_id)
        return JSONResponse(content={"department": department}, status_code=200)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@router.post("/add")
async def add_to_list(req: AddToListRequest, tma_user_id: int = Depends(get_tma_user_id)):
    """Додати товар до списку.
    req.user_id is ignored; TMA identity is used to prevent IDOR.
    """
    try:
        print(f"➕ Add to list: tma_user_id={tma_user_id}, product_id={req.product_id}, quantity={req.quantity}")
        await orm_add_item_to_temp_list(user_id=tma_user_id, product_id=req.product_id, quantity=req.quantity)
        print(f"✅ Successfully added to temp list")
        return JSONResponse(content={"success": True, "message": f"Додано {req.quantity} шт."}, status_code=200)
    except ValueError as e:
        # Помилка валідації відділу
        print(f"⚠️ Validation error: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=400)
    except Exception as e:
        print(f"❌ ERROR in add_to_list: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка додавання", "details": str(e)}, status_code=500)


@router.post("/update")
async def update_item_quantity(req: UpdateQuantityRequest, tma_user_id: int = Depends(get_tma_user_id)):
    """Оновити кількість товару.
    req.user_id is ignored; TMA identity is used to prevent IDOR.
    """
    try:
        if req.quantity < 1:
            return JSONResponse(content={"success": False, "message": "Кількість має бути більше 0"}, status_code=400)
        await orm_update_temp_list_item_quantity(user_id=tma_user_id, product_id=req.product_id, new_quantity=req.quantity)
        return JSONResponse(content={"success": True, "message": f"Кількість оновлено: {req.quantity} шт."}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in update_item_quantity: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка оновлення", "details": str(e)}, status_code=500)


@router.post("/delete")
async def delete_item(req: DeleteItemRequest, tma_user_id: int = Depends(get_tma_user_id)):
    """Видалити товар зі списку.
    req.user_id is ignored; TMA identity is used to prevent IDOR.
    """
    try:
        await orm_delete_temp_list_item(user_id=tma_user_id, product_id=req.product_id)
        return JSONResponse(content={"success": True, "message": "Товар видалено"}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in delete_item: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка видалення", "details": str(e)}, status_code=500)


@router.post("/clear/{user_id}")
async def clear_list(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """Очистити список.
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        await orm_clear_temp_list(tma_user_id)
        return JSONResponse(content={"success": True, "message": "Список очищено"}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in clear_list: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка очищення", "details": str(e)}, status_code=500)


@router.post("/save/{user_id}")
async def save_list_to_excel(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """
    Зберегти список в Excel.
    WebApp: НЕ відправляє в Telegram, тільки зберігає в archives/active/.
    Файл доступний через вкладку "Архів".
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        print(f"💾 Save list request for tma_user_id={tma_user_id} (webapp - archive only)")
        async with async_session() as session:
            async with session.begin():
                main_list_path, surplus_list_path = await process_and_save_list(session, tma_user_id)
        if not main_list_path and not surplus_list_path:
            print(f"⚠️ List is empty for tma_user_id {tma_user_id}")
            return JSONResponse(content={"success": False, "message": "Список порожній"}, status_code=400)
        print(f"✅ Files saved: main={main_list_path}, surplus={surplus_list_path}")
        return JSONResponse(content={
            "success": True,
            "message": "✅ Список збережено!",
            "cleared": True,
            "has_main": bool(main_list_path),
            "has_surplus": bool(surplus_list_path)
        }, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in save_list_to_excel: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка збереження списку", "details": str(e)}, status_code=500)


@router.get("/archives/{user_id}")
async def get_user_archives(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """Отримати список архівних файлів користувача.
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        print(f"📁 Archives request for tma_user_id={tma_user_id}")
        archives = fetch_user_archives(tma_user_id)
        if not archives:
            return JSONResponse(content={"archives": []}, status_code=200)
        result = []
        for filename, timestamp in archives:
            is_surplus = filename.startswith("лишки_")
            result.append({
                "filename": filename,
                "date": timestamp.strftime("%d.%m.%Y %H:%M"),
                "timestamp": timestamp.isoformat(),
                "is_surplus": is_surplus,
                "type": "Лишки" if is_surplus else "Основний список"
            })
        print(f"✅ Returning {len(result)} archives")
        return JSONResponse(content={"archives": result}, status_code=200)
    except Exception as e:
        print(f"❌ ERROR in get_user_archives: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка отримання архівів", "details": str(e)}, status_code=500)


@router.get("/statistics/{user_id}")
async def get_user_statistics(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """Отримати статистику користувача: кількість списків, загальна сума, популярні відділи.
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        archives = fetch_user_archives(tma_user_id)
        
        if not archives:
            return JSONResponse(content={
                "total_lists": 0,
                "total_amount": 0.0,
                "total_items": 0,
                "popular_department": None,
                "this_month_lists": 0,
                "this_month_amount": 0.0
            }, status_code=200)
        
        total_lists = len(archives)
        total_amount = 0.0
        total_items = 0
        departments = {}
        
        # Дата місяць тому
        month_ago = datetime.now() - timedelta(days=30)
        this_month_lists = 0
        this_month_amount = 0.0
        
        for filename, timestamp in archives:
            file_path = os.path.join(ACTIVE_DIR, filename)
            if not os.path.exists(file_path):
                continue
            
            try:
                # Парсимо Excel для підрахунку
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                
                file_amount = 0.0
                file_items = 0
                
                # Рахуємо товари та суму (пропускаємо заголовок)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    if str(row[0]).strip() in ["", "К-ть артикулів:", "Зібрано на суму:"]:
                        continue
                    
                    file_items += 1
                    
                    try:
                        if len(row) > 3 and row[3]:
                            sum_value = row[3]
                            if isinstance(sum_value, str):
                                sum_value = sum_value.replace(' грн', '').replace(',', '.').strip()
                            file_amount += float(sum_value)
                        else:
                            qty = float(row[1]) if len(row) > 1 and row[1] else 0
                            price = float(row[2]) if len(row) > 2 and row[2] else 0
                            file_amount += qty * price
                    except (ValueError, TypeError, IndexError):
                        pass
                
                wb.close()
                
                total_amount += file_amount
                total_items += file_items
                
                if timestamp >= month_ago:
                    this_month_lists += 1
                    this_month_amount += file_amount
                
                parsed = parse_filename(filename)
                if parsed and "department" in parsed:
                    dept = parsed["department"]
                    departments[dept] = departments.get(dept, 0) + 1
                    
            except Exception as e:
                print(f"⚠️ Error parsing {filename}: {e}")
                continue
        
        popular_department = max(departments, key=departments.get) if departments else None
        
        print(f"📊 Stats for tma_user_id {tma_user_id}: {total_lists} lists, {total_amount:.2f} грн, dept: {popular_department}")
        
        return JSONResponse(content={
            "total_lists": total_lists,
            "total_amount": round(total_amount, 2),
            "total_items": total_items,
            "popular_department": popular_department,
            "this_month_lists": this_month_lists,
            "this_month_amount": round(this_month_amount, 2)
        }, status_code=200)
        
    except Exception as e:
        print(f"❌ ERROR in get_user_statistics: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"error": "Помилка отримання статистики"}, status_code=500)


@router.get("/archives/download-all/{user_id}")
async def download_all_archives(user_id: int, tma_user_id: int = Depends(get_tma_user_id)):
    """Завантажити всі архіви користувача як ZIP.
    Path user_id is kept for URL compatibility but ignored; TMA identity is used.
    """
    try:
        archives = fetch_user_archives(tma_user_id)
        
        if not archives:
            raise HTTPException(status_code=404, detail="No archives found")
        
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, timestamp in archives:
                file_path = os.path.join(ACTIVE_DIR, filename)
                if os.path.exists(file_path):
                    zip_file.write(file_path, filename)
        
        zip_buffer.seek(0)
        zip_filename = f"epicservice_archives_{tma_user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        
        print(f"📦 Created ZIP with {len(archives)} files for tma_user_id {tma_user_id}")
        
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={zip_filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in download_all_archives: {type(e).__name__}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Download error")


@router.get("/archive/stats/{filename}")
async def get_archive_stats(filename: str, tma_user_id: int = Depends(get_tma_user_id)):
    """Отримати статистику з Excel файлу архіву.
    Query user_id (if present) is ignored; TMA identity is used for ownership check.
    """
    try:
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        parsed = parse_filename(filename)
        if not parsed or parsed["user_id"] != tma_user_id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        file_path = os.path.join(ACTIVE_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        items_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            if str(row[0]).strip() in ["", "К-ть артикулів:", "Зібрано на суму:"]:
                continue
            items_count += 1
        
        wb.close()
        department = parsed.get("department", "Невідомо")
        
        print(f"📊 Stats for {filename}: {items_count} items, department={department}, tma_author={tma_user_id}")
        
        return JSONResponse(content={
            "success": True,
            "items_count": items_count,
            "department": str(department),
            "author_id": tma_user_id
        }, status_code=200)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in get_archive_stats: {type(e).__name__}: {e}")
        traceback.print_exc()
        return JSONResponse(content={"success": False, "error": "Помилка читання статистики"}, status_code=500)


@router.get("/archive/download/{filename}")
async def download_archive(filename: str, user_id: int = Depends(get_current_user_id_any_auth)):
    """
    Завантажити архівний файл.
    Вимагає JWT Bearer токен або Telegram initData.
    Користувач може завантажувати тільки власні архіви.
    """
    try:
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")

        # Verify ownership: filename must belong to this user
        parsed = parse_filename(filename)
        if not parsed or parsed["user_id"] != user_id:
            raise HTTPException(status_code=403, detail="Access denied")

        file_path = os.path.join(ACTIVE_DIR, filename)

        # Defense-in-depth: verify resolved path stays within ACTIVE_DIR.
        if not is_path_within(_ACTIVE_DIR_PATH, Path(file_path)):
            raise HTTPException(status_code=400, detail="Invalid filename")

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in download_archive: {type(e).__name__}: {e}")
        raise HTTPException(status_code=500, detail="Download error")


@router.delete("/archive/delete/{filename}")
async def delete_archive(filename: str, user_id: int = Depends(get_current_user_id_any_auth)):
    """
    Видалити архівний файл.
    Вимагає JWT Bearer токен або Telegram initData.
    Користувач може видаляти тільки власні архіви.
    """
    try:
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")

        parsed = parse_filename(filename)
        if not parsed or parsed["user_id"] != user_id:
            print(f"⚠️ User {user_id} tried to delete file not owned by them: {filename}")
            raise HTTPException(status_code=403, detail="Access denied")

        file_path = os.path.join(ACTIVE_DIR, filename)

        # Defense-in-depth: verify resolved path stays within ACTIVE_DIR.
        if not is_path_within(_ACTIVE_DIR_PATH, Path(file_path)):
            raise HTTPException(status_code=400, detail="Invalid filename")

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        os.remove(file_path)

        return JSONResponse(content={"success": True, "message": "Файл видалено"}, status_code=200)

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in delete_archive: {type(e).__name__}: {e}")
        raise HTTPException(status_code=500, detail="Delete error")


# ===========================================================================
# Mobile App (Android) API endpoints — JWT Bearer token authentication
# ===========================================================================

def _get_user_id_from_token(authorization: str) -> int:
    """Extracts and validates user_id from a Bearer JWT token header."""
    from webapp.routers.auth import get_current_user
    if not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Невірний формат заголовка Authorization")
    return get_current_user(authorization[7:])


class SaveListMobileRequest(BaseModel):
    items: List[dict]  # [{"article_name": str, "quantity": int}]


@router.get("/products/search")
async def mobile_search_products(
    q: str,
    limit: int = 50,
    authorization: str = Header(...),
):
    """
    GET /api/products/search?q=...&limit=...
    Пошук товарів для мобільного додатку (JWT-автентифікація).
    """
    _get_user_id_from_token(authorization)
    try:
        all_products = await orm_find_products(q)
        items = all_products[:limit]
        return JSONResponse({
            "items": [
                {
                    "article": p.артикул,
                    "name": p.назва,
                    "quantity": p.кількість,
                    "department": p.відділ,
                    "group": p.група,
                    "reserved": p.відкладено,
                    "months_no_move": p.місяці_без_руху,
                    "price": p.ціна,
                    "total_value": p.сума_залишку,
                }
                for p in items
            ],
            "total": len(all_products),
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products/{article}")
async def mobile_get_product(
    article: str,
    authorization: str = Header(...),
):
    """
    GET /api/products/{article}
    Деталі товару за артикулом для мобільного додатку.
    """
    _get_user_id_from_token(authorization)
    try:
        async with async_session() as session:
            result = await session.execute(
                select(Product).where(Product.артикул == article, Product.активний)
            )
            product = result.scalar_one_or_none()
        if not product:
            raise HTTPException(status_code=404, detail="Товар не знайдено")
        return JSONResponse({
            "product": {
                "article": product.артикул,
                "name": product.назва,
                "quantity": product.кількість,
                "department": product.відділ,
                "group": product.група,
                "reserved": product.відкладено,
                "months_no_move": product.місяці_без_руху,
                "price": product.ціна,
                "total_value": product.сума_залишку,
            }
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/saved-lists")
async def mobile_get_saved_lists(authorization: str = Header(...)):
    """
    GET /api/saved-lists
    Отримати збережені списки поточного користувача (JWT-автентифікація).
    """
    user_id = _get_user_id_from_token(authorization)
    try:
        async with async_session() as session:
            result = await session.execute(
                select(SavedList)
                .where(SavedList.user_id == user_id)
                .order_by(SavedList.created_at.desc())
            )
            lists = result.scalars().all()
        return JSONResponse({
            "lists": [
                {
                    "id": sl.id,
                    "name": sl.file_name,
                    "file_name": sl.file_name,
                    "created_at": sl.created_at.isoformat() if sl.created_at else None,
                }
                for sl in lists
            ]
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/saved-lists")
async def mobile_save_list(
    req: SaveListMobileRequest,
    authorization: str = Header(...),
):
    """
    POST /api/saved-lists
    Зберегти список для поточного користувача (JWT-автентифікація).
    """
    user_id = _get_user_id_from_token(authorization)
    if not req.items:
        raise HTTPException(status_code=400, detail="Список порожній")
    try:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_name = f"list_{user_id}_{ts}.txt"
        async with async_session() as session:
            new_list = SavedList(
                user_id=user_id,
                file_name=file_name,
                file_path="",
                created_at=datetime.utcnow(),
            )
            session.add(new_list)
            await session.flush()
            for item in req.items:
                article = str(item.get("article_name", ""))
                qty = int(item.get("quantity", 1))
                session.add(SavedListItem(
                    list_id=new_list.id,
                    article_name=article,
                    quantity=qty,
                ))
            await session.commit()
            list_id = new_list.id
        return JSONResponse({"success": True, "id": list_id})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/saved-lists/{list_id}/download")
async def mobile_download_saved_list(
    list_id: int,
    token: str,
):
    """
    GET /api/saved-lists/{list_id}/download?token=...
    Завантажити збережений список як текстовий файл.
    """
    from webapp.routers.auth import get_current_user
    user_id = get_current_user(token)
    try:
        async with async_session() as session:
            result = await session.execute(
                select(SavedList).where(SavedList.id == list_id, SavedList.user_id == user_id)
            )
            saved_list = result.scalar_one_or_none()
            if not saved_list:
                raise HTTPException(status_code=404, detail="Список не знайдено")
            items_result = await session.execute(
                select(SavedListItem).where(SavedListItem.list_id == list_id)
            )
            items = items_result.scalars().all()

        lines = [f"{item.article_name}\t{item.quantity}" for item in items]
        content = "\n".join(lines).encode("utf-8")
        filename = saved_list.file_name or f"list_{list_id}.txt"

        return Response(
            content=content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
